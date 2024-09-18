import neighbourhood_scraper
import os
import time
from openpyxl.styles import Font, Color, Alignment, Side, Border, PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def create_workbook(workbook, save_as):
    
    worksheet = workbook.active
    worksheet.title = "Crime and Housing Data"

    worksheet.cell(row=1, column=1, value="City")
    worksheet.cell(row=1, column=2, value="State")
    worksheet.cell(row=1, column=3, value="Median Housing Cost")
    worksheet.cell(row=1, column=4, value="Crime Rate")

    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 35
    worksheet.column_dimensions['C'].width = 35
    worksheet.column_dimensions['D'].width = 35

    workbook.save(save_as)

    return worksheet

def workbook_table(workbook):

    worksheet = workbook["Crime and Housing Data"]
    for c in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(c)
        max_col_row = len([cell for cell in worksheet[col_letter] if cell.value])
    row = "A1:D" + str(max_col_row)

    tab = Table(displayName="DataSet", ref=row)
    
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    
    worksheet.add_table(tab)

def main():

    os.system('cls')
    print("* * * * * * * * * * * * * * * * * * * * * * * * *")
    print("*                                               *")
    print("*               City Scouter v0.1               *")
    print("*                                               *")
    print("* * * * * * * * * * * * * * * * * * * * * * * * *")

    print('\n\n\n\n\n\n\n\n')
    time.sleep(2)

    l = True
    while l is True:
        l = False
        print("What Would You Like To Do?")
        print("- - - - - - - - - - - - - - - - - - - - - - - - -")
        print("1. Get Both Crime and Housing Data for a City")
        print("2. Get Just Crime Data for a City")
        print("3. Get Just Housing Cost Data for a City")
        print("- - - - - - - - - - - - - - - - - - - - - - - - -")
        user = input(": ")
        get_data(user)

def format_city(c):
    city = ""
    for a in c:
        if a == ' ':
            city += "-"
        else:
            city += a
    return city

def get_state_code(state):

    states = {
        "alabama": "AL",
        "alaska": "AK",
        "arizona": "AZ",
        "arkansas": "AR",
        "california": "CA",
        "colorado": "CO",
        "connecticut": "CT",
        "delaware": "DE",
        "florida": "FL",
        "georgia": "GA",
        "hawaii": "HI",
        "idaho": "ID",
        "illinois": "IL",
        "indiana": "IN",
        "iowa": "IA",
        "kansas": "KS",
        "kentucky": "KY",
        "louisiana": "LA",
        "maine": "ME",
        "maryland": "MD",
        "massachusetts": "MA",
        "michigan": "MI",
        "minnesota": "MN",
        "mississippi": "MS",
        "missouri": "MO",
        "montana": "MT",
        "nebraska": "NE",
        "nevada": "NV",
        "new hampshire": "NH",
        "new jersey": "NJ",
        "new mexico": "NM",
        "new york": "NY",
        "north carolina": "NC",
        "north dakota": "ND",
        "ohio": "OH",
        "oklahoma": "OK",
        "oregon": "OR",
        "pennsylvania": "PA",
        "rhode island": "RI",
        "south carolina": "SC",
        "south dakota": "SD",
        "tennessee": "TN",
        "texas": "TX",
        "utah": "UT",
        "vermont": "VT",
        "virginia": "VA",
        "washington": "WA",
        "west virginia": "WV",
        "wisconsin": "WI",
        "wyoming": "WY",
        }
    
    s = states[state]
    return s

def get_data(userChoice):

    os.system('cls')
    d = {}
    print("What Would You Like To Do?")
    print("- - - - - - - - - - - - - - - - - - - - - - - - -")
    print("1. Single City, State Lookup")
    print("2. Bulk Lookup from .txt file")
    i = input(": ")

    if i == '1':
        s = input("Please Enter the State to Lookup Data for: ").lower()
        state = get_state_code(s)
        c = input("Please Enter the City to Lookup Data for: ")
        city = format_city(c)
        d[city] = state

    elif i == '2':
        os.system('cls')
        print("Please enter the name of the .txt file to read from")
        print("(*Note that the .txt file must exist in the same directory as this script")
        print("(**Note that the format must match exactly: 'city, state' with each entry being separated by a new line)")
        filename = input(": ")

        if '.txt' not in filename:
            filename += '.txt'

        with open(filename) as w:
            line = w.read().splitlines()
        
        for item in line:
            s = ""
            c = ""
            check = True

            for h in item:
                if check is True:
                    if ',' not in h:
                        c += h
                    else:
                        check = False
                else:
                    s += h
            
            s = s[1:len(s)].lower()

            state = get_state_code(s)
            city = format_city(c)
            d[city] = state

    os.system('cls')
    #save_as = input("Please enter a name to save the excel sheet as: ")

    #if ".xlsx" not in save_as:
        #save_as += ".xlsx"

    os.system("cls")
    print("Grabbing Data...")

    #workbook = Workbook()
    #worksheet = create_workbook(workbook, save_as)
    r = 2

    for entry in d:
        try:
            state = d[entry]
            city = entry

            print(f"{city}, {state}:")
            print("* * * * * * * * * * * * * * * * * * * * * * * *")

            crime = neighbourhood_scraper.get_crime(state, city)
            print(f"The Crime Index for {city}, {state} is {crime} out of 100.")
            print(f"This means that it is safer than {crime}% of the cities in the United States.")

            housing = neighbourhood_scraper.get_housing_market(state, city)
            print(f"The median home value in {city}, {state} is {housing}")

            #worksheet.cell(row=r, column=1, value=city)
            #worksheet.cell(row=r, column=2, value=state)
            #worksheet.cell(row=r, column=3, value=housing)
            #worksheet.cell(row=r, column=4, value=crime + " out of 100")

            r += 1
            #workbook.save(save_as)
            print("")

        except:
            state = d[entry]
            city = entry
            print(f"No Data Found for {city}, {state}")
            print("Please ensure you have entered a valid city and state (note that neighborhoods may not have free data available).")
    
    #workbook_table(workbook)

if __name__ == "__main__":
    main()